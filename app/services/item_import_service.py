from __future__ import annotations

import csv
import io

from openpyxl import load_workbook
from sqlalchemy.orm import Session as DbSession

from app.models.item import Item

REQUIRED_COLUMNS = ["item_id", "use_type", "item_text"]
OPTIONAL_COLUMNS = [
    "sentiment",
    "example_score_2",
    "example_score_1_ack",
    "example_score_1_con",
    "example_score_0",
    "hint_template",
    "pretraining_phase",
]


def _parse_csv(content: bytes) -> list[dict]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp949")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]


def _parse_xlsx(content: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    result = []
    for row in rows:
        if all(cell is None for cell in row):
            continue
        result.append({header[i]: row[i] for i in range(len(header)) if i < len(row)})
    return result


def parse_item_file(filename: str, content: bytes) -> list[dict]:
    if filename.lower().endswith(".xlsx"):
        return _parse_xlsx(content)
    return _parse_csv(content)


def upsert_items(db: DbSession, rows: list[dict]) -> tuple[int, list[str]]:
    upserted = 0
    errors = []

    for i, row in enumerate(rows, start=2):  # row 1 is the header
        row = {(k or "").strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}

        missing = [col for col in REQUIRED_COLUMNS if not row.get(col)]
        if missing:
            errors.append(f"{i}행: 필수 컬럼 누락 ({', '.join(missing)})")
            continue

        db.merge(
            Item(
                item_id=row["item_id"],
                use_type=row["use_type"],
                sentiment=row.get("sentiment") or None,
                item_text=row["item_text"],
                example_score_2=row.get("example_score_2") or None,
                example_score_1_ack=row.get("example_score_1_ack") or None,
                example_score_1_con=row.get("example_score_1_con") or None,
                example_score_0=row.get("example_score_0") or None,
                hint_template=row.get("hint_template") or None,
                pretraining_phase=row.get("pretraining_phase") or None,
                status="approved",
            )
        )
        upserted += 1

    db.commit()
    return upserted, errors
